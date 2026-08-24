"""Generates an Excel risk register from a completed VendorAssessment."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from .models import VendorAssessment
from .questions import QUESTIONS_BY_ID

_HEADER_FONT = Font(bold=True)


def _write_header(sheet: Worksheet, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = _HEADER_FONT


def _autosize(sheet: Worksheet, max_width: int = 60) -> None:
    for column in sheet.columns:
        longest = max((len(str(c.value)) for c in column if c.value is not None), default=0)
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(max(longest + 2, 12), max_width)


def _build_workbook(assessment: VendorAssessment) -> Workbook:
    profile = assessment.profile
    wb = Workbook()

    # --- Summary ---
    summary = wb.active
    summary.title = "Summary"
    rows = [
        ("Vendor", profile.vendor_name),
        ("Product / service", profile.product_name),
        ("Website", profile.vendor_website),
        ("Business owner", profile.business_owner),
        ("Engagement stage", profile.engagement_stage),
        ("Business use case", profile.business_use_case),
        ("", ""),
        ("Inherent risk", assessment.inherent_risk.value if assessment.inherent_risk else "Not scored"),
        ("Residual risk", assessment.residual_risk.value if assessment.residual_risk else "Not scored"),
        ("Residual rationale", assessment.residual_rationale),
        ("Verified control coverage", f"{assessment.control_coverage:.0%}"),
        ("", ""),
        ("Decision", assessment.decision.value),
        ("Recommendation", assessment.recommendation or "Pending human review"),
    ]
    for label, value in rows:
        summary.append([label, value])
    for row in summary.iter_rows(min_col=1, max_col=1):
        row[0].font = _HEADER_FONT
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 80
    for row in summary.iter_rows(min_col=2, max_col=2):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # --- Intake profile ---
    intake = wb.create_sheet("Intake")
    _write_header(intake, ["Field", "Value"])
    intake_rows = [
        ("Deployment model", profile.deployment_model),
        ("Business criticality", profile.criticality.value),
        ("Data classification", profile.data_classification.value),
        ("Data types", ", ".join(profile.data_types) or "—"),
        ("Regulatory scope", ", ".join(profile.regulatory_scope) or "—"),
        ("Approx. record volume", profile.record_volume or "—"),
        ("Used by", profile.used_by),
        ("Approx. user count", profile.user_count or "—"),
        ("Integrates with internal systems", "Yes" if profile.integrates_with_internal_systems else "No"),
        ("Integrated systems", profile.integrated_systems or "—"),
        ("AI capabilities", ", ".join(profile.ai_capabilities) or "—"),
        ("Affects decisions about people", "Yes" if profile.affects_decisions_about_people else "No"),
        ("Model hosting", profile.model_hosting or "—"),
    ]
    for label, value in intake_rows:
        intake.append([label, value])
    _autosize(intake)

    # --- Inherent risk drivers ---
    drivers = wb.create_sheet("Inherent Risk Drivers")
    _write_header(drivers, ["Factor", "Detail", "Points"])
    for d in assessment.inherent_drivers:
        drivers.append([d.factor, d.detail, d.points])
    drivers.append(["", "Total", sum(d.points for d in assessment.inherent_drivers)])
    _autosize(drivers)

    # --- Evidence ---
    evidence_sheet = wb.create_sheet("Evidence")
    _write_header(evidence_sheet, ["Domain", "Question", "Answer", "Verified", "Source document", "Page", "Supporting quote"])
    for e in assessment.evidence:
        question = QUESTIONS_BY_ID.get(e.question_id)
        evidence_sheet.append(
            [
                question.domain if question else "",
                e.question_text,
                e.answer if e.verified else "Not verified",
                "Yes" if e.verified else "No",
                e.citation.document if e.citation else "",
                e.citation.page if e.citation else "",
                e.quote or "",
            ]
        )
    _autosize(evidence_sheet, max_width=70)

    # --- Findings ---
    findings_sheet = wb.create_sheet("Findings")
    _write_header(findings_sheet, ["Severity", "Finding", "Recommended control"])
    for f in assessment.findings:
        findings_sheet.append([f.severity.value, f.summary, f.recommended_control or ""])
    _autosize(findings_sheet, max_width=70)

    # --- Required controls ---
    controls_sheet = wb.create_sheet("Required Controls")
    _write_header(controls_sheet, ["Required control"])
    for control in assessment.required_controls:
        controls_sheet.append([control])
    _autosize(controls_sheet, max_width=70)

    # --- RMF mapping ---
    rmf_sheet = wb.create_sheet("NIST AI RMF Mapping")
    _write_header(rmf_sheet, ["Function", "Category", "Note"])
    for m in assessment.rmf_mappings:
        rmf_sheet.append([m.function.value, m.category, m.note])
    _autosize(rmf_sheet, max_width=60)

    # --- Screened content (only when something was flagged) ---
    if assessment.injection_flags:
        flags_sheet = wb.create_sheet("Screened Content")
        _write_header(flags_sheet, ["Document", "Page", "Reason", "Snippet"])
        for flag in assessment.injection_flags:
            flags_sheet.append([flag.document, flag.page, flag.reason, flag.snippet])
        _autosize(flags_sheet, max_width=70)

    return wb


def build_excel_bytes(assessment: VendorAssessment) -> bytes:
    """Render the risk register to bytes, for a one-click download in the UI."""
    buffer = BytesIO()
    _build_workbook(assessment).save(buffer)
    return buffer.getvalue()


def write_excel_report(assessment: VendorAssessment, out_path: str | Path) -> Path:
    out_path = Path(out_path)
    out_path.write_bytes(build_excel_bytes(assessment))
    return out_path
